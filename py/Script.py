import openpyxl

def numToCharOrder(num):
    charOrder = ''
    while num > 0:
        rest = (num - 1) % 26
        charOrder = chr(65 + rest) + charOrder
        num = (num - rest) // 26
    return charOrder

def simulateYear(countries):
    return countries

def simulateYears(years, countries):
    for i in range(0, years):
        simulateYear(countries)

    return countries

# population => row=0-103, col=0-20526
workbook_pop = openpyxl.load_workbook("./pop_bothSex.xlsx")
# fertility => row=0-47, col=0-20526
workbook_fert = openpyxl.load_workbook("./WPP2022_FERT_F01_FERTILITY_RATES_BY_SINGLE_AGE_OF_MOTHER.xlsx")
# mortality => row
workbook_mort = openpyxl.load_workbook("./WPP2022_MORT_F01_1_DEATHS_SINGLE_AGE_BOTH_SEXES.xlsx")


# Population(age) + 2000 * Fertility(age) / Population(age) - Mortility(age)
countries = {}

for countryIndex in range(74, 20526, 72):
    countries[workbook_pop["A" + str(countryIndex)]] = {}
    countries[workbook_pop["A" + str(countryIndex)]]["population"] = []
    for year in range(0,100):
        countries[workbook_pop["D" + str(countryIndex)]]["population"].append(workbook_pop[numToCharOrder(68+year) + str(countryIndex)])
        print(countryIndex, year, countries[workbook_pop["A" + str(countryIndex)]]["population"][year])